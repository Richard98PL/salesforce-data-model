import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import subprocess
import requests
import os
import json

with open("config.json", "r") as file:
    config = json.load(file)

# Retrieve variables from the loaded JSON data
root_dir = config["root_dir"]
username = config["username"]
password = config["password"]
security_token = config["security_token"]
client_id = config["client_id"]
client_secret = config["client_secret"]
object_list = config["object_list"]
object_list = sorted(object_list)
standardValueSets = config["standardValueSets"]
project_directory = config["project_directory"]
auth_url = config["auth_url"]

data = {
    "grant_type": "password",
    "client_id": client_id,
    "client_secret": client_secret,
    "username": username,
    "password": password + security_token,
}

# Send the HTTP POST request
response = requests.post(auth_url, data=data)

# Get the access token from the response
access_token, instance_url = (
    response.json()["access_token"],
    response.json()["instance_url"],
)
os.environ["SFDX_ACCESS_TOKEN"] = access_token

package_xml_content = """<?xml version="1.0" encoding="UTF-8"?>
<Package xmlns="http://soap.sforce.com/2006/04/metadata">
    <types>
        <members>SolutionStatus</members>
        <members>ConsequenceOfFailure</members>
        <members>QuickTextChannel</members>
        <members>Salutation</members>
        <members>FiscalYearQuarterPrefix</members>
        <members>WorkOrderLineItemStatus</members>
        <members>OpportunityCompetitor</members>
        <members>ContactRole</members>
        <members>ChangeRequestStatus</members>
        <members>WorkOrderStatus</members>
        <members>AccountContactRole</members>
        <members>IncidentRelatedItemImpactType</members>
        <members>DigitalAssetStatus</members>
        <members>IncidentSubCategory</members>
        <members>AccountType</members>
        <members>LocationType</members>
        <members>IncidentPriority</members>
        <members>ProblemPriority</members>
        <members>ChangeRequestRiskLevel</members>
        <members>ProblemRelatedItemImpactType</members>
        <members>EventSubject</members>
        <members>ChangeRequestRelatedItemImpactLevel</members>
        <members>ProcessExceptionPriority</members>
        <members>FiscalYearPeriodPrefix</members>
        <members>TaskStatus</members>
        <members>CaseStatus</members>
        <members>FiscalYearQuarterName</members>
        <members>IncidentType</members>
        <members>LeadStatus</members>
        <members>SocialPostEngagementLevel</members>
        <members>ChangeRequestCategory</members>
        <members>ProblemRelatedItemImpactLevel</members>
        <members>ChangeRequestImpact</members>
        <members>MilitaryService</members>
        <members>FiscalYearPeriodName</members>
        <members>OpportunityStage</members>
        <members>OpportunityType</members>
        <members>ContactRequestStatus</members>
        <members>ChangeRequestBusinessReason</members>
        <members>QuantityUnitOfMeasure</members>
        <members>CaseType</members>
        <members>CaseReason</members>
        <members>IncidentStatus</members>
        <members>ContactPointAddressType</members>
        <members>ScorecardMetricCategory</members>
        <members>IdeaMultiCategory</members>
        <members>CaseContactRole</members>
        <members>StatusReason</members>
        <members>AccountRating</members>
        <members>ContactPointUsageType</members>
        <members>Industry</members>
        <members>ProblemStatus</members>
        <members>CardType</members>
        <members>IncidentReportedMethod</members>
        <members>TaskPriority</members>
        <members>ProblemSubCategory</members>
        <members>QuickTextCategory</members>
        <members>ChangeRequestPriority</members>
        <members>SocialPostClassification</members>
        <members>ContactRequestReason</members>
        <members>ServiceContractApprovalStatus</members>
        <members>TaskSubject</members>
        <members>AssociatedLocationType</members>
        <members>ProblemUrgency</members>
        <members>SocialPostReviewedStatus</members>
        <members>IncidentRelatedItemImpactLevel</members>
        <members>IncidentUrgency</members>
        <members>EventType</members>
        <members>WorkStepStatus</members>
        <members>AssetStatus</members>
        <members>CampaignMemberStatus</members>
        <members>TaskType</members>
        <members>LeadSource</members>
        <members>AccountOwnership</members>
        <members>EntitlementType</members>
        <members>WorkOrderPriority</members>
        <members>RoleInTerritory2</members>
        <members>ProblemImpact</members>
        <members>PartnerRole</members>
        <members>ContractStatus</members>
        <members>ProcessExceptionStatus</members>
        <members>CaseOrigin</members>
        <members>CampaignType</members>
        <members>ProblemCategory</members>
        <members>IncidentCategory</members>
        <members>IdeaStatus</members>
        <members>ProcessExceptionSeverity</members>
        <members>CampaignStatus</members>
        <members>CasePriority</members>
        <members>UnitOfMeasure</members>
        <members>ContractContactRole</members>
        <members>IncidentImpact</members>
        <members>Product2Family</members>
        <members>AssetRelationshipType</members>
        <members>OrderStatus</members>
        <members>OrderType</members>
        <members>ProcessExceptionCategory</members>
        <name>StandardValueSet</name>
    </types>
    <types>
        <members>*</members>
        <members>Account</members>
        <members>Contact</members>
        <members>Lead</members>
        <members>Opportunity</members>
        <members>Case</members>
        <members>Task</members>
        <members>Event</members>
        <members>Campaign</members>
        <members>Product2</members>
        <members>Pricebook2</members>
        <name>CustomObject</name>
    </types>
    <types>
        <members>*</members>
        <name>GlobalValueSet</name>
    </types>
    <version>55.0</version>
</Package>
"""

os.makedirs(project_directory, exist_ok=True)

# Create sfdx-project.json file
sfdx_project_content = """{
  "packageDirectories": [
    {
      "path": "force-app",
      "default": true
    }
  ],
  "namespace": "",
  "sfdcLoginUrl": "https://login.salesforce.com",
  "sourceApiVersion": "55.0"
}"""

sfdx_project_file = os.path.join(project_directory, "sfdx-project.json")

with open(sfdx_project_file, "w") as file:
    file.write(sfdx_project_content)

# Create force-app directory
force_app_directory = os.path.join(project_directory, "force-app")
os.makedirs(force_app_directory, exist_ok=True)

package_xml_file = os.path.join(force_app_directory, "package.xml")

if not os.path.isfile(package_xml_file):
    with open(package_xml_file, "w") as file:
        file.write(package_xml_content)

os.chdir(project_directory)

login_cmd = (
    f"sfdx org login access-token --no-prompt -r {instance_url} -a temporary_org"
)
subprocess.run(login_cmd, shell=True)

retrieve_cmd = (
    f"sfdx project retrieve start -c -o temporary_org --manifest force-app/package.xml"
)
subprocess.run(retrieve_cmd, shell=True)
os.chdir("..")
os.chdir("..")


objectByStandardValueSet = {}
for key, values in standardValueSets.items():
    for value in values:
        objectByStandardValueSet[value] = key


gray_fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
yellow_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def extract_field_info(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()

    api_name_elem = root.find("{http://soap.sforce.com/2006/04/metadata}fullName")
    api_name = api_name_elem.text if api_name_elem is not None else ""

    label_elem = root.find("{http://soap.sforce.com/2006/04/metadata}label")
    label = label_elem.text if label_elem is not None else ""

    length_elem = root.find("{http://soap.sforce.com/2006/04/metadata}length")
    length = length_elem.text if length_elem is not None else ""

    field_type_elem = root.find("{http://soap.sforce.com/2006/04/metadata}type")
    field_type = field_type_elem.text if field_type_elem is not None else ""

    unique_elem = root.find("{http://soap.sforce.com/2006/04/metadata}unique")
    unique = unique_elem.text if unique_elem is not None else ""

    precision_elem = root.find("{http://soap.sforce.com/2006/04/metadata}precision")
    precision = precision_elem.text if precision_elem is not None else ""

    scale_elem = root.find("{http://soap.sforce.com/2006/04/metadata}scale")
    scale = scale_elem.text if scale_elem is not None else ""

    required_elem = root.find("{http://soap.sforce.com/2006/04/metadata}required")
    required = required_elem.text if required_elem is not None else ""

    externalId_elem = root.find("{http://soap.sforce.com/2006/04/metadata}externalId")
    externalId = externalId_elem.text if externalId_elem is not None else ""

    value_set_elem = root.find("{http://soap.sforce.com/2006/04/metadata}valueSet")
    value_set_restricted = "false"
    value_set_values = ""

    if value_set_elem is not None:
        restricted_elem = value_set_elem.find(
            "{http://soap.sforce.com/2006/04/metadata}restricted"
        )
        if restricted_elem is not None:
            value_set_restricted = restricted_elem.text

        value_set_def_elem = value_set_elem.find(
            "{http://soap.sforce.com/2006/04/metadata}valueSetDefinition"
        )
        if value_set_def_elem is not None:
            value_elems = value_set_def_elem.findall(
                "{http://soap.sforce.com/2006/04/metadata}value"
            )
            value_set_values = "\n".join(
                [
                    ""
                    + value.find(
                        "{http://soap.sforce.com/2006/04/metadata}fullName"
                    ).text
                    for value in value_elems
                ]
            )
        else:
            value_set_name = root.find(
                ".//{http://soap.sforce.com/2006/04/metadata}valueSet/{http://soap.sforce.com/2006/04/metadata}valueSetName"
            )
            value = root.find(
                ".//{http://soap.sforce.com/2006/04/metadata}valueSet/{http://soap.sforce.com/2006/04/metadata}value"
            )

            if value_set_name is not None:
                value = value_set_name.text

            if value is not None:
                file_name = value + ".globalValueSet-meta.xml"
                file_path = os.path.join(
                    project_directory, "force-app/main/default/globalValueSets"
                )
                file_path = os.path.join(file_path, file_name)

                if os.path.exists(file_path):
                    tree = ET.parse(file_path)
                    root = tree.getroot()

                    full_names = [
                        custom_value.find(
                            "{http://soap.sforce.com/2006/04/metadata}fullName"
                        ).text
                        for custom_value in root.findall(
                            ".//{http://soap.sforce.com/2006/04/metadata}customValue"
                        )
                    ]

                    value_set_values = "\n".join(full_names)
                    field_type += ' - Global Value Set [' + value + ']'
                        

    if field_type == "Lookup":
        reference_to_elem = root.find(
            "{http://soap.sforce.com/2006/04/metadata}referenceTo"
        )
        reference_to = reference_to_elem.text if reference_to_elem is not None else ""
        field_type = f"{field_type} ({reference_to})"

    if api_name.endswith('Id') and field_type == 'Lookup ()':
        field_type = 'Lookup (' + api_name[:-2] + ')'
    
    return (
        api_name,
        label,
        length,
        field_type,
        unique,
        precision,
        scale,
        required,
        externalId,
        value_set_values,
        value_set_restricted,
    )


def get_fullnames_from_xml(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()

    # Extract the full names of the standard values
    full_names = []
    namespace = {"ns": "http://soap.sforce.com/2006/04/metadata"}
    result = ""
    for standard_value in root.findall("ns:standardValue", namespace):
        full_name = standard_value.find("ns:fullName", namespace).text
        full_names.append(full_name)

    for name in full_names:
        result += "\n" + name

    return result


def create_excel_table(objects):
    excel_file = "data_model.xlsx"

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    columns = [
        "Object",
        "Comment",
        "API Name",
        "Label",
        "Length",
        "Type",
        "Unique",
        "Precision",
        "Scale",
        "Required",
        "External Id",
        "Value Set",
        "Value Set Restricted",
        "User Story",
    ]
    all_sheet = wb.create_sheet(title="All")
    all_sheet.append(columns)

    for object_name in objects:
        object_dir = os.path.join(root_dir, "objects")
        object_dir = os.path.join(object_dir, object_name)
        fields_dir = os.path.join(object_dir, "fields")
        if not os.path.exists(fields_dir):
            print(f"Fields directory not found for object: {object_name}")
            continue

        title = object_name[:30]
        sheet = wb.create_sheet(title=title)
        sheet.append(columns)
        sheet.freeze_panes = "B2"  # Freeze the first row

        row_count = 0
        for field_file in os.listdir(fields_dir):
            if field_file.startswith("."):
                continue

            field_name = os.path.splitext(field_file)[0]
            field_path = os.path.join(fields_dir, field_file)
            (
                api_name,
                label,
                length,
                field_type,
                unique,
                precision,
                scale,
                required,
                externalId,
                value_set_values,
                value_set_restricted,
            ) = extract_field_info(field_path)
            if (
                field_type == "Picklist" and len(value_set_values) == 0
            ):  # standard picklist skip
                continue
            sheet.append(
                [
                    object_name,
                    "",
                    api_name,
                    label,
                    length,
                    field_type,
                    unique,
                    precision,
                    scale,
                    required,
                    externalId,
                    value_set_values,
                    value_set_restricted,
                    "",
                ]
            )

            row_count += 1
            for cell in sheet[row_count]:
                cell.border = thin_border

            if row_count % 2 == 1:
                for row in sheet.iter_rows(min_row=row_count, max_row=row_count):
                    for cell in row:
                        cell.fill = gray_fill

            for row in sheet.iter_rows(min_row=0, max_row=1):
                for cell in row:
                    cell.fill = yellow_fill

        last_row = len(sheet["A"])
        for row in sheet.iter_rows(min_row=last_row, max_row=last_row):
            for cell in row:
                cell.border = thin_border

        if last_row % 2 == 1:
            for row in sheet.iter_rows(min_row=last_row, max_row=last_row):
                for cell in row:
                    cell.fill = gray_fill

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                        # Adjust max_length for the "Value Set" column based on the longest substring
                        if "\n" in str(cell.value):
                            max_length = max(
                                [
                                    len(substring)
                                    for substring in str(cell.value).split("\n")
                                ]
                            )
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        sheet.auto_filter.ref = sheet.dimensions

        for value_set_column in sheet:
            for cell in value_set_column:
                cell.alignment = Alignment(wrapText=True)

        standard_value_sets_dir = os.path.join(root_dir, "standardValueSets")
        if os.path.exists(standard_value_sets_dir):
            for file_name in os.listdir(standard_value_sets_dir):
                if file_name.endswith(".standardValueSet-meta.xml"):
                    file_path = os.path.join(standard_value_sets_dir, file_name)
                    word = file_name.split(".")[0]

                    if word not in objectByStandardValueSet:
                        continue

                    object = objectByStandardValueSet[word]

                    if object not in object_list:
                        continue

                    if object_name != object:
                        continue

                    # Create a row in the appropriate sheet based on object name
                    standard_sheet_name = object[:30]

                    if standard_sheet_name not in wb.sheetnames:
                        standard_sheet = wb.create_sheet(title=standard_sheet_name)
                        standard_sheet.append(columns)
                        standard_sheet.freeze_panes = "B2"  # Freeze the first row
                    else:
                        standard_sheet = wb[standard_sheet_name]

                    # Extract information from the file
                    fullNames = get_fullnames_from_xml(file_path)

                    # Create the row with the extracted information
                    row = [
                        object,
                        "",
                        word,
                        word,
                        "",
                        "Standard Picklist",
                        "",
                        "",
                        "",
                        "",
                        "",
                        fullNames,
                        "false",
                        "",
                    ]
                    standard_sheet.append(row)

                    # Apply formatting to the row
                    row_count = len(standard_sheet["A"])
                    for cell in standard_sheet[row_count]:
                        cell.border = thin_border
                    if row_count % 2 == 1:
                        for row in standard_sheet.iter_rows(
                            min_row=row_count, max_row=row_count
                        ):
                            for cell in row:
                                cell.fill = gray_fill

                    for row in standard_sheet.iter_rows(min_row=0, max_row=1):
                        for cell in row:
                            cell.fill = yellow_fill

                    # Adjust column widths
                    for column in standard_sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                                    # Adjust max_length for the "Value Set" column based on the longest substring
                                    if "\n" in str(cell.value):
                                        max_length = max(
                                            [
                                                len(substring)
                                                for substring in str(cell.value).split(
                                                    "\n"
                                                )
                                            ]
                                        )
                            except TypeError:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        standard_sheet.column_dimensions[
                            column_letter
                        ].width = adjusted_width

                        for value_set_column in standard_sheet:
                            for cell in value_set_column:
                                cell.alignment = Alignment(wrapText=True)

        # Merge sheets into "All" sheet
        row_count = 0
        for row in sheet.iter_rows(min_row=2):
            all_sheet.append([cell.value for cell in row])

    # styles for All sheet

    for row in all_sheet.iter_rows(min_row=0, max_row=1):
        for cell in row:
            cell.fill = yellow_fill
    last_row = 0
    for row in all_sheet.iter_rows(min_row=0, max_row=len(all_sheet["A"])):
        last_row += 1
        for cell in row:
            cell.border = thin_border
            if last_row % 2 == 1 and last_row != 1:
                cell.fill = gray_fill

    # Adjust column widths in the "All" sheet
    for column in all_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        wasNewLine = False
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
                    # Adjust max_length for the "Value Set" column based on the longest substring
                    if "\n" in str(cell.value):
                        max_length = max(
                            [
                                len(substring)
                                for substring in str(cell.value).split("\n")
                            ]
                        )
                        wasNewLine = True
            except TypeError:
                pass
        parameter = 1.2
        if wasNewLine:
            parameter = 1.8
        adjusted_width = (max_length + 2) * parameter
        all_sheet.column_dimensions[column_letter].width = adjusted_width

    all_sheet.auto_filter.ref = all_sheet.dimensions

    # Set word wrap for the "Value Set" column in the "All" sheet
    for value_set_column_all in all_sheet:
        for cell in value_set_column_all:
            cell.alignment = Alignment(wrapText=True)

    # Freeze the first row and set it to bold in all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.freeze_panes = "B2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    wb.save(excel_file)
    print(f"Excel file '{excel_file}' created successfully.")


create_excel_table(object_list)



import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd

def paint_rows_white(sheet):
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

def color_alternate_rows(sheet):
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = gray_fill

# Load the Excel file
workbook = openpyxl.load_workbook('data_model.xlsx')

# Iterate through each sheet
for sheet_name in workbook.sheetnames:
    current_sheet = workbook[sheet_name]

    # Convert the sheet to a pandas DataFrame
    df = pd.DataFrame(current_sheet.values)
    header = df.iloc[0]
    df = df[1:]
    df.columns = header

    # Create a new column with lowercase 'API Name' values for sorting
    df['API Name Lower'] = df['API Name'].str.lower()

    # Sort the DataFrame by the lowercase 'API Name' column
    df.sort_values('API Name Lower', inplace=True)

    # Paint all rows white
    paint_rows_white(current_sheet)

    # Remove the lowercase 'API Name' column from the DataFrame
    df.drop('API Name Lower', axis=1, inplace=True)

    # Write the sorted DataFrame to the worksheet, preserving original case
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            current_sheet.cell(row=r_idx, column=c_idx).value = value

    # Color alternate rows gray
    color_alternate_rows(current_sheet)

    # Apply autofilter
    current_sheet.auto_filter.ref = current_sheet.dimensions

# Save the modified Excel file
workbook.save('data_model.xlsx')
